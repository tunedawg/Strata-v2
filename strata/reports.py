import io
import re
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
from flask import send_file
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .config import EXPORTS_DIR
from .storage import dataset_dir


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _generated_at() -> str:
    return datetime.now().strftime("%B %d, %Y at %I:%M %p")


def workbook_response(workbook, file_name: str):
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=file_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def build_hit_report(results: dict, dataset_name: str):
    generated_at = _generated_at()
    context_cache: dict[tuple[str, int], dict] = {}
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for term, payload in results.items():
        sheet = workbook.create_sheet(_safe_sheet_title(term))
        _write_hit_report_title(sheet, term, dataset_name, generated_at)
        headers = ["Search Term", "Document Name", "Result Page", "Bates Number (Bottom Right)", "Confidential Stamp", "Matched Text Snippet"]
        widths = [34, 54, 12, 30, 20, 100]
        header_row = 5
        _format_header(sheet, headers, widths, row=header_row)
        row_index = header_row + 1
        for row in _hit_rows(term, payload, dataset_name, context_cache):
            values = [row["term"], row["document"], row["page_label"], row["bates"], row["confidential"], row["snippet"]]
            for col_index, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_index, column=col_index, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = _thin_border()
            row_index += 1
        if row_index == header_row + 1:
            sheet.cell(row=row_index, column=1, value="No results for this term.")
        sheet.freeze_panes = "A6"
        sheet.auto_filter.ref = f"A{header_row}:F{max(row_index - 1, header_row)}"
    return workbook_response(workbook, f"strata_hit_report_{_timestamp()}.xlsx")


def build_hit_report_pdf(results: dict, dataset_name: str):
    generated_at = _generated_at()
    context_cache: dict[tuple[str, int], dict] = {}
    buffer = io.BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
    styles = getSampleStyleSheet()
    story = []
    for term, payload in results.items():
        story.append(Paragraph(f"Hit Report for {term} within {dataset_name}", styles["Title"]))
        story.append(Paragraph(f"Generated: {generated_at}", styles["Normal"]))
        story.append(Spacer(1, 10))
        table_rows = [[
            Paragraph("Document Name", styles["BodyText"]),
            Paragraph("Result Page", styles["BodyText"]),
            Paragraph("Bates Number<br/>(Bottom Right)", styles["BodyText"]),
            Paragraph("Confidential<br/>Stamp", styles["BodyText"]),
            Paragraph("Matched Text Snippet", styles["BodyText"]),
        ]]
        rows = _hit_rows(term, payload, dataset_name, context_cache)
        if rows:
            for row in rows:
                table_rows.append(
                    [
                        Paragraph(_escape_pdf(row["document"]), styles["BodyText"]),
                        Paragraph(_escape_pdf(str(row["page_label"])), styles["BodyText"]),
                        Paragraph(_escape_pdf(row["bates"]), styles["BodyText"]),
                        Paragraph(_escape_pdf(row["confidential"]), styles["BodyText"]),
                        Paragraph(_escape_pdf(row["snippet"]), styles["BodyText"]),
                    ]
                )
        else:
            table_rows.append([Paragraph("No results for this term.", styles["BodyText"]), "", "", ""])
        table = Table(table_rows, colWidths=[170, 45, 95, 75, 350], repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#FFFFFF"), colors.HexColor("#F8FAFC")]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 18))
    document.build(story)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f"strata_hit_report_{_timestamp()}.pdf", mimetype="application/pdf")


def build_production_export(dataset_root: Path, dataset_name: str, selections: list[dict]):
    timestamp = _timestamp()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Responsiveness Log"
    headers = ["Dataset", "Search Term", "Document", "Matching Labels", "Selected", "Exported At"]
    widths = [24, 36, 48, 48, 12, 26]
    _format_header(sheet, headers, widths)
    row_index = 2
    for item in selections:
        values = [
            dataset_name,
            item.get("term", ""),
            item.get("name", ""),
            ", ".join(item.get("labels", [])),
            "Yes" if item.get("selected", True) else "No",
            timestamp,
        ]
        for col_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)
        row_index += 1

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        workbook_buffer = io.BytesIO()
        workbook.save(workbook_buffer)
        workbook_buffer.seek(0)
        archive.writestr("responsiveness_log.xlsx", workbook_buffer.read())
        for item in selections:
            relative_name = item.get("name", "")
            source = dataset_root / relative_name
            if source.exists():
                archive.write(source, arcname=f"files/{relative_name.replace(chr(92), '/')}")
    zip_buffer.seek(0)
    return send_file(zip_buffer, as_attachment=True, download_name=f"strata_production_{timestamp}.zip", mimetype="application/zip")


def build_redaction_findings_xlsx(results: list[dict]):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Findings"
    headers = ["Document", "Page", "Type", "Details"]
    widths = [52, 10, 28, 80]
    _format_header(sheet, headers, widths)
    row_index = 2
    for result in results:
        for finding in result.get("findings", []):
            values = [result["filename"], finding.get("page", "-"), finding["type"], finding["details"]]
            for col_index, value in enumerate(values, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)
            row_index += 1
    return workbook_response(workbook, f"strata_redaction_findings_{_timestamp()}.xlsx")


def build_redaction_summary_xlsx(summary: dict, results: list[dict]):
    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    _format_header(summary_sheet, ["Files Scanned", "With Redactions", "Total Findings", "Clean Files"], [18, 18, 18, 18])
    for column_index, key in enumerate(["files_scanned", "with_redactions", "total_findings", "clean_files"], start=1):
        summary_sheet.cell(row=2, column=column_index, value=summary.get(key, 0))

    details_sheet = workbook.create_sheet("Documents")
    _format_header(details_sheet, ["Document", "Count", "Pages", "Types"], [52, 10, 24, 64])
    row_index = 2
    for result in results:
        values = [
            result["filename"],
            result["finding_count"],
            ", ".join(f"p.{page}" for page in result.get("pages", [])),
            "; ".join(f"{key} ({value})" for key, value in result.get("finding_types", {}).items()),
        ]
        for col_index, value in enumerate(values, start=1):
            details_sheet.cell(row=row_index, column=col_index, value=value)
        row_index += 1
    return workbook_response(workbook, f"strata_redaction_summary_{_timestamp()}.xlsx")


def build_redaction_pdf(summary: dict, results: list[dict]):
    buffer = io.BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
    styles = getSampleStyleSheet()
    story = [Paragraph("Strata Redaction Report", styles["Title"]), Spacer(1, 12)]
    story.append(
        Paragraph(
            f"Files scanned: {summary['files_scanned']} | With redactions: {summary['with_redactions']} | "
            f"Total findings: {summary['total_findings']} | Clean files: {summary['clean_files']}",
            styles["BodyText"],
        )
    )
    story.append(Spacer(1, 12))
    rows = [["Document", "Count", "Pages", "Types"]]
    for result in results:
        rows.append(
            [
                result["filename"],
                str(result["finding_count"]),
                ", ".join(f"p.{page}" for page in result.get("pages", [])),
                "; ".join(f"{key} ({value})" for key, value in result.get("finding_types", {}).items()),
            ]
        )
    table = Table(rows, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7C2D12")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#334155")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FFF7ED")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(table)
    document.build(story)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f"strata_redaction_report_{_timestamp()}.pdf", mimetype="application/pdf")


def build_redacted_zip(dataset_root: Path, results: list[dict]):
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        for result in results:
            source = dataset_root / result["filename"]
            if source.exists():
                archive.write(source, arcname=f"redacted_docs/{result['filename'].replace(chr(92), '/')}")
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f"strata_redacted_docs_{_timestamp()}.zip", mimetype="application/zip")


def _hit_rows(term: str, payload: dict, dataset_name: str, context_cache: dict[tuple[str, int], dict]) -> list[dict]:
    rows = []
    for document in payload.get("documents", []):
        document_name = document.get("name", "")
        for match in document.get("matches", []):
            page = match.get("page")
            page_number = int(page) if isinstance(page, int) or (isinstance(page, str) and page.isdigit()) else None
            page_context = _page_context(dataset_name, document_name, page_number, context_cache)
            rows.append(
                {
                    "term": term,
                    "document": document_name,
                    "page_label": _page_label(page_number),
                    "bates": page_context.get("bates", ""),
                    "confidential": page_context.get("confidential", ""),
                    "snippet": match.get("snippet", ""),
                }
            )
    return rows


def _page_label(page: int | None) -> str:
    return f"p. {page}" if page else ""


def _page_context(dataset_name: str, document_name: str, page: int | None, cache: dict[tuple[str, int], dict]) -> dict:
    if not page:
        return {"bates": "", "confidential": ""}
    key = (document_name, page)
    if key in cache:
        return cache[key]
    result = {"bates": "", "confidential": ""}
    path = dataset_dir(dataset_name) / document_name
    if path.suffix.lower() != ".pdf" or not path.exists():
        cache[key] = result
        return result
    try:
        import pdfplumber

        with pdfplumber.open(str(path)) as pdf:
            if page < 1 or page > len(pdf.pages):
                cache[key] = result
                return result
            pdf_page = pdf.pages[page - 1]
            text = pdf_page.extract_text() or ""
            result["confidential"] = "Yes" if "CONFIDENTIAL" in text.upper() else ""
            width = float(pdf_page.width)
            height = float(pdf_page.height)
            bottom_right = pdf_page.crop((width * 0.55, height * 0.78, width, height))
            candidates = []
            for word in bottom_right.extract_words() or []:
                value = (word.get("text") or "").strip()
                if _looks_like_bates(value):
                    candidates.append(value)
            result["bates"] = candidates[-1] if candidates else ""
    except Exception:
        pass
    cache[key] = result
    return result


def _looks_like_bates(value: str) -> bool:
    normalized = value.strip().strip(".,;:()[]{}")
    if len(normalized) < 4 or len(normalized) > 40:
        return False
    if not any(char.isdigit() for char in normalized):
        return False
    if re.fullmatch(r"\d{1,3}", normalized):
        return False
    return bool(re.fullmatch(r"[A-Za-z]{0,12}[-_ ]?\d{4,12}[A-Za-z0-9_-]*", normalized))


def _safe_sheet_title(term: str) -> str:
    cleaned = "".join("_" if char in "[]:*?/\\" else char for char in str(term)).strip()
    return (cleaned or "Hit Report")[:31]


def _thin_border() -> Border:
    side = Side(style="thin", color="CBD5E1")
    return Border(left=side, right=side, top=side, bottom=side)


def _write_hit_report_title(sheet, term: str, dataset_name: str, generated_at: str) -> None:
    sheet.merge_cells("A1:F1")
    title = sheet["A1"]
    title.value = f"Hit Report for {term} within {dataset_name}"
    title.font = Font(bold=True, size=16, color="111827")
    title.alignment = Alignment(horizontal="left")
    sheet.merge_cells("A2:F2")
    generated = sheet["A2"]
    generated.value = f"Generated: {generated_at}"
    generated.font = Font(size=10, color="475569")
    generated.alignment = Alignment(horizontal="left")
    sheet.merge_cells("A3:F3")
    note = sheet["A3"]
    note.value = "Each row below is one search result. Bates and confidentiality values are detected from the matched PDF page when available."
    note.font = Font(italic=True, size=10, color="64748B")


def _escape_pdf(value: str) -> str:
    return str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _format_header(sheet, headers: list[str], widths: list[int], row: int = 1) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="111827")
    for index, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = sheet.cell(row=row, column=index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = _thin_border()
        sheet.column_dimensions[cell.column_letter].width = width
